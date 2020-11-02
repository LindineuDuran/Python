import sys
import os
import pandas as pd
import openpyxl
from datetime import datetime

from PyQt5 import uic, QtWidgets as qtw
from PyQt5.QtGui import QIcon
from PyQt5 import QtCore as qtc
from PyQt5.QtCore import Qt as qt, QPersistentModelIndex, QModelIndex

class Ui_MainWindow(qtw.QMainWindow):
    application_path = ''

    # Configura interface
    def __init__(self, parent=None):
        super(qtw.QMainWindow, self).__init__()
        dirname = os.path.dirname(os.path.abspath(__file__))
        uic.loadUi(os.path.join(dirname,'MultiPushButtonsDeleteQTable.ui'), self)

        """Guarda o Diretório da Aplicação"""
        self.application_path = os.getcwd()

        """buttons"""
        self.btnProcess.clicked.connect(self.processar)
        self.btnSave.clicked.connect(self.savefile)
        self.btnExit.clicked.connect(qtw.QApplication.quit)

    # Rotina principal de tratamento de dados
    def processar(self):
        """Executa Processamento Escolhido"""
        # path_to_file = os.path.join(self.application_path, 'TestePandas.txt')
        path_to_file = os.path.join(self.application_path, 'DadosNFSe.txt')

        # Read the CSV into a pandas data frame (df) With a df you can do many things most important: visualize data with Seaborn
        df = pd.read_csv(path_to_file, delimiter=';', encoding='latin-1')

        header_labels = df.columns

        self.table.setColumnCount(df.shape[1]+1)
        self.table.setRowCount(df.shape[0])
        self.table.setHorizontalHeaderLabels(header_labels.insert(0, '#'))

        for index in range(df.shape[0]):
            aqr_icon = os.path.join(self.application_path, 'images', 'cross.png')
            self.btn_del = qtw.QPushButton(QIcon(aqr_icon),'')
            self.btn_del.clicked.connect(self.handleButtonClicked)
            self.table.setCellWidget(index,0,self.btn_del)

            for col in range(df.shape[1]):
                item = qtw.QTableWidgetItem(str(df.iloc[index, col]))
                self.table.setItem(index,col+1,item)

        """Ajusta primeira coluna"""
        header = self.table.horizontalHeader()
        # header.setSectionResizeMode(0, qtw.QHeaderView.Stretch)
        header.setSectionResizeMode(0, qtw.QHeaderView.ResizeToContents)

    def handleButtonClicked(self):
        # button = QtGui.qApp.focusWidget()
        button = self.sender()
        if button:
            row = self.table.indexAt(button.pos()).row()
            self.table.removeRow(row)

    def savefile(self):
        filename = qtw.QFileDialog.getSaveFileName(self, 'Save File', '', ".xls(*.xls)")
        wb = openpyxl.Workbook()
        wb.create_sheet()

        sheet = wb['Sheet']
        sheet.title = 'Teste'

        model = self.table.model()
        for c in range(model.columnCount()):
            teext = model.headerData(c, qt.Horizontal)
            sheet.cell(row=1, column=c+1).value = teext

        for currentColumn in range(self.table.columnCount()):
            for currentRow in range(self.table.rowCount()):
                if self.table.item(currentRow, currentColumn) is not None:
                    teext = str(self.table.item(currentRow, currentColumn).text())
                else:
                    teext = ''

                sheet.cell(row=currentRow+2, column=currentColumn+1).value = teext

        wb.save(filename[0])

if __name__ == "__main__":
    app = qtw.QApplication(sys.argv)
    window = Ui_MainWindow()
    window.show()
    sys.exit(app.exec_())
